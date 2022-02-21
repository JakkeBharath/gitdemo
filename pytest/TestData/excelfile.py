import openpyxl

BOOK=openpyxl.load_workbook("C:\\Users\\hpadmin\\Documents\\PYTHONDEMO.xlsx")
sheet=BOOK.active
Dict={}
cell=sheet.cell(row=1,column=1)
print(cell.value)
sheet.cell(row=2,column=2).value="BharathKumar"
#print(sheet.cell(row=2,column=2).value)
cell=sheet.cell(row=2,column=2)
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)
sheet['A6'].value="jakke"
print(sheet.cell(row=6,column=1).value)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TC3":
        for j in range (2,sheet.max_column+1):

            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
            #print(sheet.cell(row=i,column=j).value)
print(Dict)





