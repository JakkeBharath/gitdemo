import openpyxl
class HomePageData:

    Input_Data = [{"FirstName":"Jakke Bharath Kumar","E-Mail ID":"jakkebharath@gmail.com","Password":"Bharath@@2022","Gender":"Male"}]


    @staticmethod
    def getTestData(Test_Case_Name):
        Dict={}
        BOOK = openpyxl.load_workbook("C:\\Users\\hpadmin\\Documents\\PYTHONDEMO.xlsx")
        sheet = BOOK.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "Test_Case_Name":
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    #print(sheet.cell(row=i,column=j).value)

        return [Dict]


